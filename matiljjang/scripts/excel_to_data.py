#!/usr/bin/env python3
"""
맛일짱 식당 데이터 변환기
엑셀(맛일짱-식당데이터.xlsx) → src/data/restaurants.js 자동 생성

사용법:
    python3 scripts/excel_to_data.py [엑셀경로]

엑셀경로를 생략하면 저장소 루트의 '맛일짱-식당데이터.xlsx'를 사용합니다.
"""
import sys, os, json
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
DEFAULT_XLSX = os.path.join(REPO_ROOT, "맛일짱-식당데이터.xlsx")
OUT_JS = os.path.join(HERE, "..", "src", "data", "restaurants.js")

def to_bool(v):
    return str(v).strip().upper() in ("TRUE", "1", "Y", "O", "예", "있음")

def num(v, cast=float):
    if v is None or v == "":
        return None
    try:
        n = cast(v)
        return int(n) if cast is float and n == int(n) else n
    except (ValueError, TypeError):
        return v

def rows_as_dicts(ws):
    """1행=한글라벨, 2행=영문키, 3행~=데이터"""
    keys = [c.value for c in ws[2]]
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        out.append({k: v for k, v in zip(keys, row) if k})
    return out

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx):
        print(f"[오류] 엑셀 파일을 찾을 수 없어요: {xlsx}")
        sys.exit(1)

    wb = load_workbook(xlsx, data_only=True)
    sheet = {s.split(" ")[-1]: s for s in wb.sheetnames}  # '🍽️ 식당목록' -> '식당목록'
    ws_rest = wb[sheet["식당목록"]]
    ws_menu = wb[sheet["메뉴"]]
    ws_rev  = wb[sheet["리뷰"]]

    rests = rows_as_dicts(ws_rest)
    menus = rows_as_dicts(ws_menu)
    revs  = rows_as_dicts(ws_rev)

    # group menus / reviews by restaurantId
    menu_by = {}
    for m in menus:
        rid = num(m.get("restaurantId"), int)
        menu_by.setdefault(rid, []).append({
            "name": m.get("name", "") or "",
            "price": num(m.get("price")) or 0,
            "tag": m.get("tag", "") or "",
            "desc": m.get("desc", "") or "",
            "emoji": m.get("emoji", "") or "🍽️",
        })
    rev_by = {}
    for rv in revs:
        rid = num(rv.get("restaurantId"), int)
        tags = rv.get("tags", "") or ""
        rev_by.setdefault(rid, []).append({
            "author": rv.get("author", "") or "익명",
            "when": rv.get("when", "") or "최근",
            "rating": num(rv.get("rating"), int) or 5,
            "tags": [t.strip() for t in str(tags).split(",") if t.strip()],
            "text": rv.get("text", "") or "",
        })

    restaurants = []
    for r in rests:
        rid = num(r.get("id"), int)
        if rid is None:
            continue
        obj = {
            "id": rid,
            "name": r.get("name", "") or "",
            "category": r.get("category", "") or "한식",
            "tag": r.get("tag", "") or "",
            "emoji": r.get("emoji", "") or "🍽️",
            "priceMin": num(r.get("priceMin")) or 0,
            "priceMax": num(r.get("priceMax")) or 0,
            "rating": num(r.get("rating")) or 0,
            "reviewCount": num(r.get("reviewCount"), int) or 0,
            "distance": num(r.get("distance"), int) or 0,
            "busy": r.get("busy", "") or "보통",
            "hasSolo": to_bool(r.get("hasSolo")),
            "cookTime": num(r.get("cookTime"), int) or 0,
            "address": r.get("address", "") or "",
            "hours": r.get("hours", "") or "",
            "breakTime": r.get("breakTime", "") or "",
            "mapX": num(r.get("mapX"), int) or 0,
            "mapY": num(r.get("mapY"), int) or 0,
            "menus": menu_by.get(rid, []),
            "seedReviews": rev_by.get(rid, []),
        }
        lat, lng = num(r.get("lat")), num(r.get("lng"))
        if lat not in (None, "") and lng not in (None, ""):
            obj["lat"] = lat
            obj["lng"] = lng
        restaurants.append(obj)

    body = json.dumps(restaurants, ensure_ascii=False, indent=2)
    js = "export const RESTAURANTS = " + body + "\n"
    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write(js)

    print(f"변환 완료 → {os.path.relpath(OUT_JS, REPO_ROOT)}")
    print(f"  식당 {len(restaurants)}개 / 메뉴 {sum(len(o['menus']) for o in restaurants)}개 / 리뷰 {sum(len(o['seedReviews']) for o in restaurants)}개")
    geo = sum(1 for o in restaurants if 'lat' in o)
    print(f"  좌표(lat/lng) 입력된 식당: {geo}개")

if __name__ == "__main__":
    main()
